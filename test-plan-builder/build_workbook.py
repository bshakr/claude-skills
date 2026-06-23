#!/usr/bin/env python
"""Test-plan workbook generator.

Reusable template for the `test-plan-builder` skill. Produces a multi-tab .xlsx:
  README  ·  RISK REGISTER & E2E  ·  one tab per research domain.

Each scenario row: ID, Area, Scenario, Priority, Preconditions, Steps,
Expected result, Edge/Risk, Code ref, + blank Status (dropdown) / Tester / Notes.

HOW TO USE
  1. python3 -m venv .venv && ./.venv/bin/pip install openpyxl
  2. Fill TITLE, README_*, RISKS, FLOWS, and the SHEETS list from your agent reports
     (keep the file:line refs verbatim — they make failures triageable).
  3. ./.venv/bin/python build_workbook.py
The sample data below is illustrative — replace it.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ====================== FILL THESE IN ======================
OUT = "Test_Plan.xlsx"
TITLE = "<Feature> — Test Plan"
README_SUBTITLE = ("Test plan derived from the ACTUAL implemented code, not the tickets. "
                   "Generated against latest default branches: <repo@sha>, ...")
README_GATES = [
    "FILL: launch scope / feature flags / allowed values / permission model.",
]
README_CAVEATS = [  # implementation-state reality vs spec (built / partial / not-built)
    "FILL: surfaces that are NOT built or only partial — so QA doesn't write impossible cases.",
]
RISKS = [  # (tag, text) — lead with these
    ("AREA Xn", "FILL: top risk a tester must weight heavily (silent data loss, no-op delete, dead-end, missing isolation)."),
]
FLOWS = [  # ordered end-to-end happy paths to run first
    "E2E-1 FILL: a full cross-layer journey, from setup to an observable outcome.",
]

# Each sheet: (tab_title, intro_line, [rows]); each row is a 9-tuple:
# (ID, Area, Scenario, Priority, Preconditions, Steps, Expected, Edge/Risk, Code ref)
SHEETS = [
    ("DOMAIN-1 Example", "FILL: one-line intro: routes/gates/key invariants for this domain.", [
        ("X1", "Area", "Scenario title", "P0",
         "Preconditions", "1. step\n2. step", "Expected observable result",
         "Edge / risk note", "path/to/file.py:123"),
    ]),
]
# ===========================================================

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F3864")
SUB_FONT = Font(italic=True, size=10, color="555555")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
P_FILL = {"P0": PatternFill("solid", fgColor="F4B6B6"),
          "P1": PatternFill("solid", fgColor="FBE2A0"),
          "P2": PatternFill("solid", fgColor="DDE6F0")}
ALT_FILL = PatternFill("solid", fgColor="F5F7FB")
RISK_FILL = PatternFill("solid", fgColor="FFD9CC")

COLS = ["ID", "Area", "Scenario", "Pri", "Preconditions", "Steps",
        "Expected result", "Edge / Risk", "Code ref", "Status", "Tester", "Notes / Defect"]
WIDTHS = [7, 20, 30, 5, 34, 38, 44, 38, 30, 11, 10, 22]


def add_sheet(wb, name, intro, rows):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    c = ws.cell(1, 1, intro); c.font = SUB_FONT; c.alignment = WRAP
    ws.row_dimensions[1].height = 42
    for j, col in enumerate(COLS, 1):
        cell = ws.cell(2, j, col)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
    for j, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A3"
    r = 3
    for row in rows:
        vals = list(row) + ["", "", ""]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(r, j, v)
            cell.alignment = CENTER if j in (1, 4, 10) else WRAP
            cell.border = BORDER
            if r % 2 == 1:
                cell.fill = ALT_FILL
        pcell = ws.cell(r, 4)
        pcell.fill = P_FILL.get(row[3], P_FILL["P2"]); pcell.font = Font(bold=True)
        r += 1
    dv = DataValidation(type="list", formula1='"Not run,Pass,Fail,Blocked,N/A,Deferred"', allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"J3:J{r-1}")
    return ws


def text_sheet(wb, name, blocks):
    """blocks: list of (text, kind) where kind in {title, sub, head, body, risk}."""
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 120
    r = 2
    for text, kind in blocks:
        c = ws.cell(r, 2, text); c.alignment = WRAP
        if kind == "title": c.font = TITLE_FONT; ws.row_dimensions[r].height = 24
        elif kind == "sub": c.font = SUB_FONT; ws.row_dimensions[r].height = 46
        elif kind == "head": c.font = Font(bold=True, size=12, color="1F3864")
        elif kind == "risk":
            c.fill = RISK_FILL; c.border = BORDER; ws.row_dimensions[r].height = 44
        else: ws.row_dimensions[r].height = 30
        r += 1
    return ws


wb = openpyxl.Workbook(); wb.remove(wb.active)

readme = [(TITLE, "title"), (README_SUBTITLE, "sub"), ("LAUNCH SCOPE / GATES", "head")]
readme += [(g, "body") for g in README_GATES]
readme += [("IMPLEMENTATION-STATE CAVEATS — do NOT write impossible cases", "head")]
readme += [(g, "risk") for g in README_CAVEATS]
readme += [("HOW TO USE", "head"),
           ("Priority: P0 = launch-blocking, P1 = important, P2 = edge. Status is a dropdown. "
            "Code ref points at the implementation backing each case. Run the RISK REGISTER & E2E tab first.", "body")]
text_sheet(wb, "README", readme)

risk = [("Top Risks — weight these heavily before go-live", "title")]
risk += [(f"[{tag}]  {txt}", "risk") for tag, txt in RISKS]
risk += [("Suggested end-to-end happy-path flows (run first)", "head")]
risk += [(f, "body") for f in FLOWS]
text_sheet(wb, "RISK REGISTER & E2E", risk)

for name, intro, rows in SHEETS:
    add_sheet(wb, name, intro, rows)

# README + risk first, domain tabs after (creation order already does this)
total = sum(len(rows) for _, _, rows in SHEETS)
wb.save(OUT)
print("WROTE", OUT, "| domains:", len(SHEETS), "| scenarios:", total)
